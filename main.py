#This Python file is for "search.html". It is repl of "Database" in Replit
import openpyxl as xl

print('On what basis do you want to perform search of allergens?\nWrite "A" for Searching on the basis of Family Type\nWrite "B" for Searching on the basis of Source')
try:
  searchType = input().lower()

  def read_Data(name_of_excel_file,name_of_excel_sheet):
    wb = xl.load_workbook(name_of_excel_file)
    sheet = wb[name_of_excel_sheet]
    return sheet

  def gettingData(sheet,column):
    possibilities = []
    for row in range(2,sheet.max_row+1):
      data = str(sheet.cell(row, column).value)
      possibilities.append(data)
    l = []
    for i in possibilities:
      if i not in l:
        l.append(i)
    print('\nWrite exactly from one of these options')
    for i in range(len(l)):
      print('If you want to search allergens in',l[i],', type "'+str(i)+'"')
    index = int(input())
    check = l[index]
    print('\n')
    return check
    

  def giveData(sheet,column,check):
    i = 0
    for row in range(2,sheet.max_row+1):
      data = str(sheet.cell(row, column).value)
      ProteinName = str(sheet.cell(row,3).value)
      AllergenName = str(sheet.cell(row, 5).value)
      NumberOfResidues = str(sheet.cell(row,6).value)
      Structure = str(sheet.cell(row, 7).value)
      if check==data:
        i = i+1
        print(str(i),'. Protein Name:',ProteinName,'\nAllergen Name:',AllergenName,'\nNumber of Residues:',NumberOfResidues,'\nStructure:',Structure,'\n')

  sheet = read_Data('Database.xlsx', 'Form responses 1')
  column = 0
  if searchType=='a':
    column = 9
  if searchType=='b':
    column = 10
  if column==0:
    print('Please type either "A" or "B"')
  check = gettingData(sheet,column)
  giveData(sheet,column,check)
  
except:
  print('Opps, something went wrong')
