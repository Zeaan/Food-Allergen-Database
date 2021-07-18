import openpyxl as xl

def read_Data(name_of_excel_file,name_of_excel_sheet):
    wb = xl.load_workbook(name_of_excel_file)
    sheet = wb[name_of_excel_sheet]
    return sheet

def HTML_Commands(sheet):
    i = 0
    for row in range(2,sheet.max_row+1):
        i = i + 1
        ProteinName = str(sheet.cell(row,2).value)
        SequenceLink = str(sheet.cell(row,3).value)
        AllergenName = str(sheet.cell(row, 4).value)
        NumberOfResidues = str(sheet.cell(row,5).value)
        Structure = str(sheet.cell(row, 6).value)
        StructureLink = str(sheet.cell(row, 7).value)
        FamilyType = str(sheet.cell(row, 8).value)
        Source = str(sheet.cell(row, 9).value)
        print("<tr>")
        print('<th scope = "row"> '+str(i)+'</th>')
        print('<td> <a href="'+SequenceLink+'" target="_blank">' + ProteinName + '</a></th>')
        print('<td> ' + AllergenName + '</th>')
        print('<td> ' + NumberOfResidues + '</th>')
        print('<td> <a href="'+StructureLink+'" target="_blank">' + Structure + '</a></th>')
        print('<td> ' + FamilyType + '</th>')
        print('<td> ' + Source + '</th>')
        print("</tr>")

name_of_excel_file = input("Enter name of the excel file ")
name_of_excel_sheet = input("Enter name of the sheet in excel file ")
sheet = read_Data(name_of_excel_file, name_of_excel_sheet)
HTML_Commands(sheet)
